from tkinter import messagebox
import rarfile
import zipfile
import re
import pandas as pd
from urllib.parse import urlparse
import tkinter as tk
from tkinter import Button
from tkinter import filedialog
from datetime import datetime
from openpyxl import Workbook
import os



current_date = datetime.now().date()
excel_path = 'database/' + str(current_date) + '.xlsx'
excel_path_2 = 'database/' + str(current_date) + '-vjp.xlsx'
excel_path_3 = 'database/' + str(current_date) + '-sdt.xlsx'
rar_path = ''
password_list = []


def read_rar(rar_path):
    results = []
    with rarfile.RarFile(rar_path) as rf:
        for item in rf.infolist():
            if item.is_file():
                if "Passwords.txt" in item.filename or "passwords.txt" in item.filename:
                    if "VN[" in item.filename or '/VN' in item.filename or 'VN_' in item.filename or '[VN]' in item.filename:
                        print(item.filename)
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern1 = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern2 = re.compile(r'url: .+?\nlogin: .+?\npassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern3 = re.compile(r'URL: .+?\nUSER: .+?\nPASS: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches1 = pattern1.findall(content)
                            matches2 = pattern2.findall(content)
                            matches3 = pattern3.findall(content)
                            # Chuyển đổi các chuỗi từ pattern 2 và pattern 3 về format chuẩn như pattern 1
                            matches2 = [match.replace('url:', 'URL:').replace('login:', 'Username:').replace('password:', 'Password:') for match in matches2]
                            matches3 = [match.replace('USER:', 'Username:').replace('PASS:', 'Password:') for match in matches3]
                            for match in matches1:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches2:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches3:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                    elif str(item.filename).startswith("VN"):
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern1 = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern2 = re.compile(r'url: .+?\nlogin: .+?\npassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern3 = re.compile(r'URL: .+?\nUSER: .+?\nPASS: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches1 = pattern1.findall(content)
                            matches2 = pattern2.findall(content)
                            matches3 = pattern3.findall(content)
                            print("matches3")
                            # Chuyển đổi các chuỗi từ pattern 2 và pattern 3 về format chuẩn như pattern 1
                            matches2 = [match.replace('url:', 'URL:').replace('login:', 'Username:').replace('password:', 'Password:') for match in matches2]
                            matches3 = [match.replace('USER:', 'Username:').replace('PASS:', 'Password:') for match in matches3]
                            for match in matches1:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches2:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches3:
                                lines = match.strip().split('\n')
                                results.extend(lines)
    cleaned_results = [item.rstrip('\r') for item in results]
    return cleaned_results

def read_rar_have_password(rar_path, password):
    results = []
    with rarfile.RarFile(rar_path) as rf:
        rf.setpassword(password)
        for item in rf.infolist():
            if item.is_file():
                if "Passwords.txt" in item.filename or "passwords.txt" in item.filename:
                    if "VN[" in item.filename or '/VN' in item.filename or 'VN_' in item.filename or '[VN]' in item.filename:
                        print(item.filename)
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern1 = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL | re.IGNORECASE)
                            pattern2 = re.compile(r'url: .+?\nlogin: .+?\npassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL | re.IGNORECASE)
                            pattern3 = re.compile(r'URL: .+?\nUSER: .+?\nPASS: .+?(?=\n|$)', re.MULTILINE | re.DOTALL | re.IGNORECASE)
                            matches1 = pattern1.findall(content)
                            matches2 = pattern2.findall(content)
                            matches3 = pattern3.findall(content)
                            # Chuyển đổi các chuỗi từ pattern 2 và pattern 3 về format chuẩn như pattern 1
                            matches2 = [match.replace('url:', 'URL:').replace('login:', 'Username:').replace('password:', 'Password:') for match in matches2]
                            matches3 = [match.replace('USER:', 'Username:').replace('PASS:', 'Password:') for match in matches3]
                            for match in matches1:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches2:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches3:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                    elif str(item.filename).startswith("VN"):
                        with rf.open(item.filename) as file:
                            content = file.read( ).decode('utf-8')
                            pattern1 = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern2 = re.compile(r'url: .+?\nlogin: .+?\npassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            pattern3 = re.compile(r'URL: .+?\nUSER: .+?\nPASS: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches1 = pattern1.findall(content)
                            matches2 = pattern2.findall(content)
                            matches3 = pattern3.findall(content)
                            # Chuyển đổi các chuỗi từ pattern 2 và pattern 3 về format chuẩn như pattern 1
                            matches2 = [match.replace('url:', 'URL:').replace('login:', 'Username:').replace('password:', 'Password:') for match in matches2]
                            matches3 = [match.replace('USER:', 'Username:').replace('PASS:', 'Password:') for match in matches3]
                            
                            for match in matches1:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches2:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            for match in matches3:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            #results.extend(matches)
                            #results = '\n'.join(match.strip() for match in matches)        
    cleaned_results = [item.rstrip('\r') for item in results]
    return cleaned_results
 
def read_zip(rar_path):
    results = []
    with zipfile.ZipFile(rar_path) as rf:
        for item in rf.infolist():
            if item.is_file():
                if "Passwords.txt" in item.filename or "passwords.txt" in item.filename:
                    if "VN[" in item.filename or '/VN' in item.filename or 'VN_' in item.filename or '[VN]' in item.filename:
                        print(item.filename)
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches = pattern.findall(content)
                            for match in matches:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                    elif str(item.filename).startswith("VN"):
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches = pattern.findall(content)
                            for match in matches:
                                lines = match.strip().split('\n')
                                results.extend(lines)
    cleaned_results = [item.rstrip('\r') for item in results]
    return cleaned_results

def read_zip_have_password(rar_path, password):
    results = []
    with zipfile.ZipFile(rar_path) as rf:
        rf.setpassword(password)
        for item in rf.infolist():
            if item.is_file():
                if "Passwords.txt" in item.filename or "passwords.txt" in item.filename:
                    if "VN[" in item.filename or '/VN' in item.filename or 'VN_' in item.filename or '[VN]' in item.filename or '-VN' in item.filename:
                        print(item.filename)
                        with rf.open(item.filename) as file:
                            content = file.read().decode('utf-8')
                            pattern = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches = pattern.findall(content)
                            for match in matches:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                    elif str(item.filename).startswith("VN"):
                        with rf.open(item.filename) as file:
                            content = file.read( ).decode('utf-8')
                            pattern = re.compile(r'URL: .+?\nUsername: .+?\nPassword: .+?(?=\n|$)', re.MULTILINE | re.DOTALL)
                            matches = pattern.findall(content)
                            for match in matches:
                                lines = match.strip().split('\n')
                                results.extend(lines)
                            #results.extend(matches)
                            #results = '\n'.join(match.strip() for match in matches)        
    cleaned_results = [item.rstrip('\r') for item in results]
    return cleaned_results

def get_main_domain(url):
    url = str(url).replace('www.','')
    parsed_url = urlparse(url)
    main_domain = str(parsed_url.netloc)
    main_domain.replace('vi-vn.facebook.com','facebook.com')
    main_domain.replace('m.facebook.com','facebook.com')
    return main_domain

def convert_from_txt_to_dataframe(lines):
    data = {'URL': [], 'Username': [], 'Password': []}
    for i in range(0, len(lines)):
        if (i+2) < len(lines):
            if ("URL:" in str(lines[i])) and ('Username:' in str(lines[i+1])) and ('Password:' in str(lines[i+2])):
                data['URL'].append(get_main_domain(lines[i].replace('URL: ','')))
                data['Username'].append(str(lines[i + 1].replace('Username: ','')).replace('\r', ''))
                data['Password'].append(lines[i + 2].replace('Password: ',''))
    df_a = pd.DataFrame(data)
    return df_a

def clean_dataframe(df):
    df = df[df['URL'].apply(lambda x: len(x) <= 40)]
    df = df.drop_duplicates()
    def check_row(row):
        if row.astype(str).str.contains('UNKNOW').any() or row.isna().any() or (row == '').any():
            return False
        return True
    df = df[df.apply(check_row, axis=1)]
    df = df.reset_index(drop=True)
    return df

def append_data_to_excel(excel_path, new_data):
    create_output()
    df_xlsx = pd.read_excel(excel_path)
    new_data = pd.DataFrame(new_data)
    if len(new_data) > 0:
        cols = ['URL', 'Username', 'Password']
        df_xlsx_ = pd.concat([df_xlsx, new_data[cols]], ignore_index=True)   
        open(excel_path, 'w').close()
        df_xlsx_.to_excel(excel_path, index=False, sheet_name="Sheet1")
    return len(new_data)

def open_file_dialog(text):
    collected = 0
    collected_vjp = 0
    rar_path = filedialog.askopenfilename()
    password = str(text.get("1.0", "end-1c"))
    if rar_have_password(rar_path):
        if str(password) == "":
            messagebox.showwarning("Alert", "enter password")
        else:
            if check_password(rar_path,password):
                results = read_rar_have_password(rar_path, password)
                df_a = convert_from_txt_to_dataframe(results)
                df_a = clean_dataframe(df_a)
                df_vjp = vjp_df(df_a)
                collected = append_data_to_excel(excel_path, df_a)
                collected_vjp = append_data_to_excel(excel_path_2, df_vjp)
            else:
                messagebox.showwarning("Alert", "wrong password")
    else:
        results = read_rar(rar_path)
        df_a = convert_from_txt_to_dataframe(results)
        df_a = clean_dataframe(df_a)       
        df_vjp = vjp_df(df_a)
        collected = append_data_to_excel(excel_path, df_a)
        collected_vjp = append_data_to_excel(excel_path_2, df_vjp)
    messagebox.showwarning("Alert", "Collected " + str(collected) + ", " + str(collected_vjp) + " vjp account! ")
    messagebox.showwarning("Alert", "Collected " + str(collected) + " vjp account! ")

def rar_have_password(file_path):
    try:
        rf = rarfile.RarFile(file_path)
        for f in rf.infolist():
            if f.needs_password():
                return True
        return False
    except rarfile.BadRarFile:
        print("Not valid file")
        return False
    except Exception as e:
        print(f"Error: {e}")
        return False

def check_password(rar_path, password):
    try:
        with rarfile.RarFile(rar_path, 'r') as rf:
            rf.setpassword(password) 
            if rf.infolist():
                first_file_info = rf.infolist()[0]
                rf.extract(first_file_info, pwd=password)
                return True
            else:
                print("RAR is Empty")
                return False
    except rarfile.BadRarFile:
        print("BadRarFile")
        return False
    except rarfile.BadPassword:
        print("Wrong Pass")
        return False
    except rarfile.Error as e:
        print(f"Unidentify: {e}")
        return False
    
def check_password_without_extract(rar_path, password):
    try:
        with rarfile.RarFile(rar_path, 'r') as rf:
            rf.setpassword(password)
            try:
                first_file_info = rf.infolist()[0]
                with rf.open(first_file_info) as f:
                    f.read(1)  
                return True
            except rarfile.BadRarFile:
                print("BadRarFile")
                return False
            except rarfile.BadPassword:
                print("Wrong Pass")
                return False
            except rarfile.Error as e:
                print(f"Unidentify: {e}")
                return False
    except rarfile.Error as e:
        print(f"Error opening file: {e}")
        return False

def write_list_to_file(lst, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:
        for item in lst:
            file.write(f"{item}\n")

def unique_domain(df):
    df_url = df['URL']
    df_url = df_url.drop_duplicates()
    filtered_df = df_url[df_url.apply(lambda x: len(x) <= 40)]
    filtered_df = filtered_df.reset_index(drop=True)
    return filtered_df

def filter_df(df, keyword_list):
    regex_pattern = '|'.join(keyword_list)
    df_filtered = df[df['URL'].str.contains(regex_pattern, na=False)]
    return df_filtered

def create_output():
    exists_1 = os.path.exists(excel_path)
    exists_2 = os.path.exists(excel_path_2)
    if exists_1:
        if exists_2:
            return
        else:
            wb = Workbook()
            ws = wb.active
            wb.save(excel_path_2) 
            return
    else:
        wb = Workbook()
        ws = wb.active
        wb.save(excel_path)
        if exists_2:
            return
        else:
            wb = Workbook()
            ws = wb.active
            wb.save(excel_path_2) 
            return

def create_file_selector_window():
    root = tk.Tk()
    root.title("File Selector")
    button_select_file = tk.Button(root, text="Select File", command=lambda: open_file_dialog(text))
    button_select_file.grid(column=0, row=0, padx=10, pady=10)
    #button_start = tk.Button(root, text="Select File", command=lambda: open_file_dialog(text))
    text = tk.Text(root, width=20, height=10, font=('Arial', 14))
    text.grid(column=0, row=1, padx=10, pady=10)
    root.mainloop()
#create_file_selector_window() 



def vjp_df(df_a):
    if len(df_a) > 1:
        keywords = ['canva','netflix', 'aws', 'paypal', 'azure', 'digitalocean.com', 'oracle','matbao', 'glint', 'work', 'lancer', 'inance', 'azdigi', 'hosting', 'amazon', '.gov.vn']
        regex_pattern = '|'.join(keywords)
        df_a = pd.DataFrame(df_a)
        df_vjp = df_a[df_a['URL'].str.contains(regex_pattern, case=False)]
        print(df_vjp)
        return df_vjp

def core_process():
    check, password = check_password_list_without_extract(rar_path, password_list)
    print("File: " + rar_path)
    print("Password đúng: " + str(password))
    if rar_have_password(rar_path):
        if str(password) == "":
            messagebox.showwarning("Alert", "enter password")
        else:
            if check_password_without_extract(rar_path,password):
                results = read_rar_have_password(rar_path, password)
                df_a = convert_from_txt_to_dataframe(results)
                df_a = clean_dataframe(df_a)
                df_vjp = vjp_df(df_a)
                collected = append_data_to_excel(excel_path, df_a)
                collected_vjp = append_data_to_excel(excel_path_2, df_vjp)
            else:
                messagebox.showwarning("Alert", "wrong password")
    else:
        results = read_rar(rar_path)
        df_a = convert_from_txt_to_dataframe(results)
        df_a = clean_dataframe(df_a)
        df_vjp = vjp_df(df_a)       
        collected = append_data_to_excel(excel_path, df_a)
        collected_vjp = append_data_to_excel(excel_path_2, df_vjp)
    messagebox.showwarning("Alert", "Collected " + str(collected) + " account, " + str(collected_vjp) + " account vjp!")

def select_file():
    global rar_path
    rar_path = filedialog.askopenfilename()
    return rar_path


def start_button(text):
    global password_list
    password_str = str(text.get("1.0", "end-1c"))
    password_list = password_str.splitlines()
    core_process()


def check_password_list(rar_path, password_list):
    for password in password_list:
        try:
            with rarfile.RarFile(rar_path, 'r') as rf:
                rf.setpassword(password) 
                if rf.infolist():
                    first_file_info = rf.infolist()[0]
                    rf.extract(first_file_info, pwd=password)
                    return 1, password
        except :
            print("Password sai")
            next
    return 0, ''

def check_password_list_without_extract(rar_path, password_list):
    for password in password_list:
        try:
            with rarfile.RarFile(rar_path, 'r') as rf:
                rf.setpassword(password)
                first_file_info = rf.infolist()[0]
                with rf.open(first_file_info) as f:
                    f.read(1)  
                return 1, password
        except rarfile.Error as e:
            print(f"Lỗi khác: {e}")
            return 0, ''
    return 0, ''  


def filter_and_save_excel(file_path, file_path_2, keywords):
    try:
        df = pd.read_excel(file_path)

        if not all(col in df.columns for col in ['URL', 'Username', 'Password']):
            print("File không chứa đầy đủ các cột URL, Username, Password.")
            return

        def contains_keyword(value):
            if pd.isna(value):
                return False, None
            for keyword in keywords:
                if keyword in str(value):
                    return True, keyword
            return False, None

        def has_consecutive_digits(value):
            if pd.isna(value):
                return False
            return bool(re.search(r'\d{9,10}', str(value)))

        matched_rows = []

        for index, row in df.iterrows():
            username_contains, username_keyword = contains_keyword(row['Username'])
            password_contains, password_keyword = contains_keyword(row['Password'])

            username_has_digits = has_consecutive_digits(row['Username'])
            password_has_digits = has_consecutive_digits(row['Password'])

            if (username_contains and password_has_digits) or (password_contains and username_has_digits):
                matched_rows.append({
                    'URL': row['URL'],
                    'Username': row['Username'],
                    'Password': row['Password'],
                    'Matched Keyword': username_keyword if username_contains else password_keyword
                })

        filtered_df = pd.DataFrame(matched_rows)

        if not filtered_df.empty:
            print("Các dòng khớp với yêu cầu:")
            print(filtered_df[['URL', 'Username', 'Password', 'Matched Keyword']])

            filtered_df.to_excel(file_path_2, index=False)
            print(f"Dữ liệu đã được ghi vào file: {file_path_2}")
        else:
            print("Không có dòng nào khớp với yêu cầu.")

    except Exception as e:
        print(f"Có lỗi xảy ra: {e}")

def create_file_selector_window_2():
    root = tk.Tk()
    root.title("File Selector")
    button_select_file = tk.Button(root, text="Select File", command=lambda: select_file())
    button_select_file.grid(column=0, row=0, padx=10, pady=10)
    button_start = tk.Button(root, text="Start", command=lambda: start_button(text))
    button_start.grid(column=0, row=1, padx=10, pady=10)
    text = tk.Text(root, width=20, height=10, font=('Arial', 14))
    text.grid(column=0, row=2, padx=10, pady=10)
    root.mainloop()

def test():
    rar_path = "C:\\Users\\PC\\Desktop\\Getaccount\\log\\MIRAGE CLOUD.rar"
    result = read_rar(rar_path)
    df = convert_from_txt_to_dataframe(result)
    filtered_df = unique_domain(df)
    vjp_df(filtered_df)
    

create_file_selector_window_2() 
#test()






